from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_user, login_required, logout_user, current_user
from models import db, User, Vehicle, Instrumentation, Activity
from werkzeug.utils import secure_filename
import os
from datetime import datetime, timedelta
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from email_utils import send_email
from rc_processor import process_rc_card
from excel_backup import add_vehicle_to_backup, backup_vehicles_to_excel, read_backup_data, restore_from_backup, backup_all_data, auto_update_backup

routes = Blueprint('routes', __name__)
UPLOAD_FOLDER = "static/uploads"
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def log_activity(vehicle_id, description):
    """Log activity for audit trail"""
    activity = Activity(
        vehicle_id=vehicle_id,
        description=description,
        date=datetime.now().strftime('%Y-%m-%d'),
        performed_by=current_user.username
    )
    db.session.add(activity)
    db.session.commit()


# ================= LOGIN =================
@routes.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']

        user = User.query.filter_by(username=username).first()

        if user and user.password == password and user.role == role:
            login_user(user)
            log_activity(0, f"User {username} logged in")
            return redirect('/dashboard')
        else:
            flash("Invalid credentials or role", "error")

    return render_template('login.html')


# ================= FORGOT PASSWORD =================
@routes.route('/forgot', methods=['GET', 'POST'])
def forgot():
    if request.method == 'POST':
        username = request.form.get('username')
        security_answer = request.form.get('security_answer', '').lower()
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        user = User.query.filter_by(username=username).first()
        
        if not user:
            flash("User not found", "error")
            return render_template('forgot.html')
        
        # Verify new password and confirm password match
        if new_password != confirm_password:
            flash("Passwords do not match", "error")
            return render_template('forgot.html')
        
        if len(new_password) < 6:
            flash("Password must be at least 6 characters", "error")
            return render_template('forgot.html')
        
        # For now, just verify it's an admin (can be enhanced with security questions)
        # Security answer check: "bosch" is the answer for all users (default)
        if security_answer == "bosch" or security_answer == "admin":
            user.password = new_password
            db.session.commit()
            flash("Password updated successfully! You can now login.", "success")
            return redirect('/')
        else:
            flash("Invalid security answer", "error")
            return render_template('forgot.html')

    return render_template('forgot.html')


# ================= DASHBOARD =================
@routes.route('/dashboard')
@login_required
def dashboard():
    search = request.args.get('search', '').strip().lower()

    if search:
        # Use ilike for case-insensitive search (works with SQLite and PostgreSQL)
        vehicles = Vehicle.query.filter(
            (Vehicle.name.ilike(f'%{search}%')) |
            (Vehicle.rc.ilike(f'%{search}%')) |
            (Vehicle.chassis.ilike(f'%{search}%')) |
            (Vehicle.owner.ilike(f'%{search}%'))
        ).all()
    else:
        vehicles = Vehicle.query.all()

    total = Vehicle.query.count()
    testing = Vehicle.query.filter_by(status="In Testing").count()
    test_completed = Vehicle.query.filter_by(status="Test Completed").count()
    returned = Vehicle.query.filter_by(status="Returned").count()

    return render_template(
        'dashboard.html',
        vehicles=vehicles,
        total=total,
        testing=testing,
        test_completed=test_completed,
        returned=returned,
        user=current_user
    )


# ================= ADMIN PANEL =================
@routes.route('/admin')
@login_required
def admin():
    if current_user.role != "Admin":
        flash("Access denied", "error")
        return redirect('/dashboard')
    
    # Stats
    total_users = User.query.count()
    total_vehicles = Vehicle.query.count()
    testing_count = Vehicle.query.filter_by(status='In Testing').count()
    test_completed_count = Vehicle.query.filter_by(status='Test Completed').count()
    returned_count = Vehicle.query.filter_by(status='Returned').count()
    
    # Percentages
    testing_percentage = int((testing_count / total_vehicles * 100)) if total_vehicles > 0 else 0
    test_completed_percentage = int((test_completed_count / total_vehicles * 100)) if total_vehicles > 0 else 0
    returned_percentage = int((returned_count / total_vehicles * 100)) if total_vehicles > 0 else 0
    
    # Role distribution
    role_distribution = {}
    roles = db.session.query(User.role, func.count(User.id)).group_by(User.role).all()
    for role, count in roles:
        role_distribution[role] = count
    
    # Get all users and vehicles
    users = User.query.all()
    vehicles = Vehicle.query.all()
    
    # Recent activities
    recent_activities = [
        {'description': 'System initialized', 'type': 'SYSTEM', 'created_at': 'Today'},
        {'description': 'Vehicles in testing', 'type': 'VEHICLE', 'created_at': 'Today'},
    ]
    
    # Monthly stats
    vehicles_added_month = Vehicle.query.filter(Vehicle.received_on >= datetime.now() - timedelta(days=30)).count()
    avg_testing_days = 5
    completion_rate = 85
    
    # Mail configuration status
    from flask import current_app
    mail_config = {
        'server': current_app.config.get('MAIL_SERVER', 'Not configured'),
        'port': current_app.config.get('MAIL_PORT', 'Not configured'),
        'use_tls': current_app.config.get('MAIL_USE_TLS', False),
        'username': current_app.config.get('MAIL_USERNAME', 'Not configured'),
        'sender': current_app.config.get('MAIL_DEFAULT_SENDER', 'Not configured'),
        'admin_email': current_app.config.get('ADMIN_EMAIL', 'Not configured'),
        'is_configured': bool(current_app.config.get('MAIL_USERNAME'))
    }
    
    return render_template('admin.html',
        total_users=total_users,
        total_vehicles=total_vehicles,
        testing_count=testing_count,
        test_completed_count=test_completed_count,
        returned_count=returned_count,
        testing_percentage=testing_percentage,
        test_completed_percentage=test_completed_percentage,
        returned_percentage=returned_percentage,
        role_distribution=role_distribution,
        users=users,
        vehicles=vehicles,
        recent_activities=recent_activities,
        vehicles_added_month=vehicles_added_month,
        avg_testing_days=avg_testing_days,
        completion_rate=completion_rate,
        mail_config=mail_config
    )


# ================= RECEIVE VEHICLE =================
@routes.route('/receive', methods=['GET', 'POST'])
@login_required
def receive():
    if current_user.role == "User":
        flash("You don't have permission to receive vehicles", "error")
        return redirect('/dashboard')
    
    if request.method == 'POST':
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        
        # Handle vehicle image
        file = request.files.get('image')
        filename = None
        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            filename = timestamp + filename
            file.save(os.path.join(UPLOAD_FOLDER, filename))
        
        # Handle RC card file
        rc_file = request.files.get('rc_card')
        rc_filename = None
        if rc_file and rc_file.filename:
            rc_filename = secure_filename(rc_file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            rc_filename = 'rc_' + timestamp + rc_filename
            rc_file.save(os.path.join(UPLOAD_FOLDER, rc_filename))
        
        # Handle insurance file
        ins_file = request.files.get('insurance_file')
        ins_filename = None
        if ins_file and ins_file.filename:
            ins_filename = secure_filename(ins_file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            ins_filename = 'insurance_' + timestamp + ins_filename
            ins_file.save(os.path.join(UPLOAD_FOLDER, ins_filename))
        
        # Handle checklist item images
        scratches_img_filename = None
        scratches_img = request.files.get('scratches_image')
        if scratches_img and scratches_img.filename and allowed_file(scratches_img.filename):
            scratches_img_filename = secure_filename(scratches_img.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            scratches_img_filename = 'scratches_' + timestamp + scratches_img_filename
            scratches_img.save(os.path.join(UPLOAD_FOLDER, scratches_img_filename))
        
        dents_img_filename = None
        dents_img = request.files.get('dents_image')
        if dents_img and dents_img.filename and allowed_file(dents_img.filename):
            dents_img_filename = secure_filename(dents_img.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            dents_img_filename = 'dents_' + timestamp + dents_img_filename
            dents_img.save(os.path.join(UPLOAD_FOLDER, dents_img_filename))
        
        glass_img_filename = None
        glass_img = request.files.get('glass_damage_image')
        if glass_img and glass_img.filename and allowed_file(glass_img.filename):
            glass_img_filename = secure_filename(glass_img.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            glass_img_filename = 'glass_' + timestamp + glass_img_filename
            glass_img.save(os.path.join(UPLOAD_FOLDER, glass_img_filename))
        
        lights_img_filename = None
        lights_img = request.files.get('lights_image')
        if lights_img and lights_img.filename and allowed_file(lights_img.filename):
            lights_img_filename = secure_filename(lights_img.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            lights_img_filename = 'lights_' + timestamp + lights_img_filename
            lights_img.save(os.path.join(UPLOAD_FOLDER, lights_img_filename))

        v = Vehicle(
            name=request.form.get('name'),
            rc=request.form.get('rc'),
            chassis=request.form.get('chassis'),
            engine=request.form.get('engine'),
            owner=request.form.get('owner'),
            received_on=request.form.get('received_on'),
            inertia=request.form.get('inertia'),
            cc=request.form.get('cc'),
            gears=request.form.get('gears'),
            road_a=request.form.get('road_a'),
            road_b=request.form.get('road_b'),
            road_c=request.form.get('road_c'),
            fuel=request.form.get('fuel'),
            oil=request.form.get('oil'),
            coolant=request.form.get('coolant'),
            start_date=request.form.get('start_date'),
            end_date=request.form.get('end_date'),
            reg_date=request.form.get('reg_date'),
            mfg_date=request.form.get('mfg_date'),
            image=filename,
            bike=request.form.get('bike'),
            insurance=request.form.get('insurance'),
            number_plate=request.form.get('number_plate'),
            rc_card=rc_filename,
            insurance_file=ins_filename,
            status="In Testing",
            # Vehicle Condition Checklist
            number_of_keys=int(request.form.get('number_of_keys', 1) or 1),
            scratches_present=bool(request.form.get('check_scratches')),
            scratches_details=request.form.get('scratches_details'),
            scratches_image=scratches_img_filename,
            dents_present=bool(request.form.get('check_dents')),
            dents_details=request.form.get('dents_details'),
            dents_image=dents_img_filename,
            glass_damage_present=bool(request.form.get('check_glass_damage')),
            glass_damage_details=request.form.get('glass_damage_details'),
            glass_damage_image=glass_img_filename,
            tire_condition=request.form.get('tire_condition'),
            battery_status=request.form.get('battery_status'),
            lights_issue_present=bool(request.form.get('check_lights_issue')),
            lights_details=request.form.get('lights_details'),
            lights_image=lights_img_filename,
            engine_status=request.form.get('engine_status'),
            condition_notes=request.form.get('condition_notes')
        )
        db.session.add(v)
        db.session.commit()
        
        # Auto-update backup Excel with all data
        auto_update_backup(db)

        inst = Instrumentation(
            vehicle_id=v.id,
            type=request.form.get('inst_type'),
            activity=request.form.get('inst_activity'),
            request_date=request.form.get('inst_request_date'),
            completion_date=request.form.get('inst_completion_date'),
            details=request.form.get('inst_details')
        )
        db.session.add(inst)
        db.session.commit()

        log_activity(v.id, f"Vehicle {v.name} received by {current_user.username}")
        
        # Send email notification
        send_email(
            subject=f"New Vehicle Received: {v.name}",
            recipients=['admin@bikevault.com'],
            template='new_vehicle',
            data={
                'vehicle_name': v.name,
                'rc_number': v.rc,
                'owner': v.owner,
                'received_on': v.received_on,
                'received_by': current_user.username
            }
        )
        
        flash(f"Vehicle {v.name} added successfully", "success")
        return redirect('/dashboard')

    return render_template('receive.html')


# ================= VEHICLE DETAILS =================
@routes.route('/vehicle/<int:id>')
@login_required
def vehicle(id):
    v = Vehicle.query.get(id)
    if not v:
        flash("Vehicle not found", "error")
        return redirect('/dashboard')
    
    inst = Instrumentation.query.filter_by(vehicle_id=id).all()
    act = Activity.query.filter_by(vehicle_id=id).order_by(Activity.timestamp.desc()).all()

    return render_template('vehicle.html', v=v, inst=inst, act=act)


# ================= EDIT VEHICLE =================
@routes.route('/edit_vehicle/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_vehicle(id):
    v = Vehicle.query.get(id)
    if not v:
        flash("Vehicle not found", "error")
        return redirect('/dashboard')
    
    if current_user.role == "User":
        flash("You don't have permission to edit vehicles", "error")
        return redirect(f'/vehicle/{id}')
    
    if request.method == 'POST':
        # Update basic info
        v.name = request.form.get('name')
        v.rc = request.form.get('rc')
        v.chassis = request.form.get('chassis')
        v.engine = request.form.get('engine')
        v.owner = request.form.get('owner')
        v.received_on = request.form.get('received_on')
        
        # Update new fields
        v.bike = request.form.get('bike')
        v.insurance = request.form.get('insurance')
        v.number_plate = request.form.get('number_plate')
        
        # Update technical specs
        v.inertia = request.form.get('inertia')
        v.cc = request.form.get('cc')
        v.gears = request.form.get('gears')
        v.road_a = request.form.get('road_a')
        v.road_b = request.form.get('road_b')
        v.road_c = request.form.get('road_c')
        v.fuel = request.form.get('fuel')
        v.oil = request.form.get('oil')
        v.coolant = request.form.get('coolant')
        v.start_date = request.form.get('start_date')
        v.end_date = request.form.get('end_date')
        v.reg_date = request.form.get('reg_date')
        v.mfg_date = request.form.get('mfg_date')
        
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        
        # Handle new image
        file = request.files.get('image')
        if file and file.filename and allowed_file(file.filename):
            # Delete old image
            if v.image and os.path.exists(os.path.join(UPLOAD_FOLDER, v.image)):
                os.remove(os.path.join(UPLOAD_FOLDER, v.image))
            
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            filename = timestamp + filename
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            v.image = filename
        
        # Handle RC card file
        rc_file = request.files.get('rc_card')
        if rc_file and rc_file.filename:
            # Delete old RC card file
            if v.rc_card and os.path.exists(os.path.join(UPLOAD_FOLDER, v.rc_card)):
                os.remove(os.path.join(UPLOAD_FOLDER, v.rc_card))
            
            rc_filename = secure_filename(rc_file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            rc_filename = 'rc_' + timestamp + rc_filename
            rc_file.save(os.path.join(UPLOAD_FOLDER, rc_filename))
            v.rc_card = rc_filename
        
        # Handle insurance file
        ins_file = request.files.get('insurance_file')
        if ins_file and ins_file.filename:
            # Delete old insurance file
            if v.insurance_file and os.path.exists(os.path.join(UPLOAD_FOLDER, v.insurance_file)):
                os.remove(os.path.join(UPLOAD_FOLDER, v.insurance_file))
            
            ins_filename = secure_filename(ins_file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            ins_filename = 'insurance_' + timestamp + ins_filename
            ins_file.save(os.path.join(UPLOAD_FOLDER, ins_filename))
            v.insurance_file = ins_filename
        
        db.session.commit()
        
        # Auto-update backup Excel with all data
        auto_update_backup(db)
        
        log_activity(v.id, f"Vehicle {v.name} updated by {current_user.username}")
        flash(f"Vehicle {v.name} updated successfully", "success")
        return redirect(f'/vehicle/{id}')

    return render_template('edit_vehicle.html', v=v)


# ================= DELETE VEHICLE =================
@routes.route('/delete_vehicle/<int:id>', methods=['POST'])
@login_required
def delete_vehicle(id):
    v = Vehicle.query.get(id)
    if not v:
        flash("Vehicle not found", "error")
        return redirect('/dashboard')
    
    if current_user.role != "Admin":
        flash("Only Admin can delete vehicles", "error")
        return redirect(f'/vehicle/{id}')
    
    # Delete image
    if v.image and os.path.exists(os.path.join(UPLOAD_FOLDER, v.image)):
        os.remove(os.path.join(UPLOAD_FOLDER, v.image))
    
    # Delete RC card file
    if v.rc_card and os.path.exists(os.path.join(UPLOAD_FOLDER, v.rc_card)):
        os.remove(os.path.join(UPLOAD_FOLDER, v.rc_card))
    
    # Delete insurance file
    if v.insurance_file and os.path.exists(os.path.join(UPLOAD_FOLDER, v.insurance_file)):
        os.remove(os.path.join(UPLOAD_FOLDER, v.insurance_file))
    
    # Delete related records
    Instrumentation.query.filter_by(vehicle_id=id).delete()
    Activity.query.filter_by(vehicle_id=id).delete()
    
    vehicle_name = v.name
    db.session.delete(v)
    db.session.commit()
    
    # Auto-update backup Excel with all data
    auto_update_backup(db)
    
    log_activity(0, f"Vehicle {vehicle_name} deleted by {current_user.username}")
    flash(f"Vehicle {vehicle_name} deleted successfully", "success")
    return redirect('/dashboard')


# ================= EXPORT TO EXCEL =================
@routes.route('/export_vehicle/<int:id>')
@login_required
def export_vehicle(id):
    v = Vehicle.query.get(id)
    if not v:
        flash("Vehicle not found", "error")
        return redirect('/dashboard')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle Details"
    
    # Define styles
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"Vehicle Details - {v.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    # Basic Information
    row = 3
    headers = ['Field', 'Value']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    data = [
        ['Vehicle Name', v.name],
        ['RC Number', v.rc],
        ['Chassis Number', v.chassis],
        ['Engine Number', v.engine or 'N/A'],
        ['Owner', v.owner or 'N/A'],
        ['Received On', v.received_on],
        ['Status', v.status],
        ['', ''],
        ['Engine CC', v.cc or 'N/A'],
        ['Gears', v.gears or 'N/A'],
        ['Inertia', v.inertia or 'N/A'],
        ['Road Load A', v.road_a or 'N/A'],
        ['Road Load B', v.road_b or 'N/A'],
        ['Road Load C', v.road_c or 'N/A'],
        ['Fuel Type', v.fuel or 'N/A'],
        ['Oil Grade', v.oil or 'N/A'],
        ['Coolant Grade', v.coolant or 'N/A'],
        ['', ''],
        ['Start Date', v.start_date or 'N/A'],
        ['End Date', v.end_date or 'N/A'],
    ]
    
    row = 4
    for item in data:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row, col, value)
            cell.border = border
        row += 1
    
    # Instrumentation
    row += 1
    ws.cell(row, 1, "INSTRUMENTATION").font = Font(bold=True, size=12)
    row += 1
    
    inst_headers = ['Type', 'Activity', 'Request Date', 'Completion Date', 'Details']
    for col, header in enumerate(inst_headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row += 1
    for i in Instrumentation.query.filter_by(vehicle_id=id).all():
        ws.cell(row, 1, i.type)
        ws.cell(row, 2, i.activity)
        ws.cell(row, 3, i.request_date or '')
        ws.cell(row, 4, i.completion_date or '')
        ws.cell(row, 5, i.details or '')
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_activity(v.id, f"Vehicle {v.name} exported by {current_user.username}")
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"{v.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


# ================= DOWNLOAD VEHICLES BACKUP =================
@routes.route('/download_backup')
@login_required
def download_backup():
    """Download all vehicles backup Excel file"""
    if current_user.role != "Admin":
        flash("Access denied. Only admins can download backups.", "error")
        return redirect('/dashboard')
    
    backup_file = "backups/vehicles_backup.xlsx"
    
    if not os.path.exists(backup_file):
        # Create backup if it doesn't exist
        backup_vehicles_to_excel(db)
    
    if not os.path.exists(backup_file):
        flash("Backup file not found", "error")
        return redirect('/dashboard')
    
    log_activity(0, f"Backup downloaded by {current_user.username}")
    
    return send_file(
        backup_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"vehicles_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


# ================= VIEW BACKUP DATA =================
@routes.route('/view_backup')
@login_required
def view_backup():
    """View all vehicles from backup Excel file"""
    if current_user.role != "Admin":
        flash("Access denied. Only admins can view backup.", "error")
        return redirect('/dashboard')
    
    backup_data = read_backup_data()
    return render_template('view_backup.html', vehicles=backup_data, total=len(backup_data))


# ================= RESTORE FROM BACKUP =================
@routes.route('/restore_backup', methods=['POST'])
@login_required
def restore_backup():
    """Restore all vehicles from backup Excel file"""
    if current_user.role != "Admin":
        flash("Access denied. Only admins can restore backup.", "error")
        return redirect('/dashboard')
    
    success, message = restore_from_backup(db)
    
    if success:
        flash(message, "success")
        log_activity(0, f"Database restored from backup by {current_user.username}")
    else:
        flash(message, "error")
    
    return redirect('/dashboard')


# ================= ADD INSTRUMENTATION =================
@routes.route('/add_inst/<int:id>', methods=['POST'])
@login_required
def add_inst(id):
    inst = Instrumentation(vehicle_id=id, **request.form)
    db.session.add(inst)
    db.session.commit()
    
    # Auto-update backup Excel with all data
    auto_update_backup(db)
    
    v = Vehicle.query.get(id)
    log_activity(id, f"Instrumentation added by {current_user.username}")
    flash("Instrumentation added successfully", "success")
    return redirect(f'/vehicle/{id}')


# ================= ADD ACTIVITY =================
@routes.route('/add_activity/<int:id>', methods=['POST'])
@login_required
def add_activity(id):
    activity = Activity(
        vehicle_id=id,
        description=request.form.get('description'),
        date=request.form.get('date'),
        performed_by=current_user.username
    )
    db.session.add(activity)
    db.session.commit()
    
    # Auto-update backup Excel with all data
    auto_update_backup(db)
    
    flash("Activity logged successfully", "success")
    return redirect(f'/vehicle/{id}')


# ================= LOGOUT =================
@routes.route('/logout')
@login_required
def logout():
    log_activity(0, f"User {current_user.username} logged out")
    logout_user()
    return redirect('/')


# ================= RETURN VEHICLE =================
@routes.route('/return_vehicle/<int:id>', methods=['POST'])
@login_required
def return_vehicle(id):
    v = Vehicle.query.get(id)
    if v:
        # Handle return image upload
        return_image = request.files.get('return_image')
        return_image_filename = None
        if return_image and return_image.filename and allowed_file(return_image.filename):
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            return_image_filename = secure_filename(return_image.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            return_image_filename = 'return_' + timestamp + return_image_filename
            return_image.save(os.path.join(UPLOAD_FOLDER, return_image_filename))
            v.return_image = return_image_filename
        
        v.status = "Returned"
        v.end_date = request.form.get('return_date')
        db.session.commit()
        log_activity(v.id, f"Vehicle {v.name} marked as returned by {current_user.username}")
        
        # Send email notification
        send_email(
            subject=f"Vehicle Returned: {v.name}",
            recipients=['admin@bikevault.com'],
            template='vehicle_returned',
            data={
                'vehicle_name': v.name,
                'rc_number': v.rc,
                'owner': v.owner,
                'return_date': v.end_date,
                'returned_by': current_user.username
            }
        )
        
        flash(f"Vehicle {v.name} marked as returned successfully", "success")
    return redirect(f'/vehicle/{id}')

# ================= CANCEL RETURN VEHICLE =================
@routes.route('/cancel_return_vehicle/<int:id>', methods=['POST'])
@login_required
def cancel_return_vehicle(id):
    v = Vehicle.query.get(id)
    if v:
        # Set status back to Test Completed (the logical prior status before Returned)
        v.status = "Test Completed"
        v.end_date = None
        db.session.commit()
        log_activity(v.id, f"Vehicle {v.name} return cancelled by {current_user.username}")
        flash(f"Vehicle {v.name} return cancelled successfully. Status reset to Test Completed", "success")
    return redirect(f'/vehicle/{id}')

# ================= MARK TEST COMPLETED =================
@routes.route('/test_completed_vehicle/<int:id>', methods=['POST'])
@login_required
def test_completed_vehicle(id):
    v = Vehicle.query.get(id)
    if v:
        v.status = "Test Completed"
        db.session.commit()
        log_activity(v.id, f"Vehicle {v.name} marked as test completed by {current_user.username}")
        
        # Send email notification
        send_email(
            subject=f"Test Completed: {v.name}",
            recipients=['admin@bikevault.com'],
            template='test_completed',
            data={
                'vehicle_name': v.name,
                'rc_number': v.rc,
                'owner': v.owner,
                'completed_by': current_user.username,
                'completion_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        )
        
        flash(f"Vehicle {v.name} marked as test completed successfully", "success")
    return redirect(f'/vehicle/{id}')

# ================= CANCEL TEST COMPLETED =================
@routes.route('/cancel_test_completed/<int:id>', methods=['POST'])
@login_required
def cancel_test_completed(id):
    v = Vehicle.query.get(id)
    if v:
        v.status = "In Testing"
        db.session.commit()
        log_activity(v.id, f"Vehicle {v.name} test completion cancelled by {current_user.username}")
        flash(f"Vehicle {v.name} status reset to In Testing", "success")
    return redirect(f'/vehicle/{id}')


# ================= TEST EMAIL ENDPOINT =================
@routes.route('/api/test_email', methods=['POST'])
@login_required
def test_email_api():
    if current_user.role != 'Admin':
        return jsonify({'success': False, 'error': 'Permission denied'}), 403
    
    try:
        admin_email = request.app.config.get('ADMIN_EMAIL', 'admin@bikevault.com')
        
        send_email(
            subject='🧪 Test Email from BikeVault',
            recipients=[admin_email],
            body_text=f'This is a test email from BikeVault System.\n\nSent by: {current_user.username}\nTime: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        )
        
        return jsonify({
            'success': True,
            'message': 'Test email sent successfully',
            'recipient': admin_email
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ================= RC CARD EXTRACTION ENDPOINT =================
@routes.route('/api/extract_rc_details', methods=['POST'])
@login_required
def extract_rc_details():
    """Extract details from RC card image using OCR"""
    try:
        # Check if RC card file was uploaded
        if 'rc_card' not in request.files:
            return jsonify({'success': False, 'error': 'No RC card image provided'}), 400
        
        rc_file = request.files['rc_card']
        
        if rc_file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Save uploaded file temporarily
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        temp_filename = secure_filename(f'temp_rc_{datetime.now().timestamp()}.jpg')
        temp_path = os.path.join(UPLOAD_FOLDER, temp_filename)
        rc_file.save(temp_path)
        
        # Process RC card image
        result = process_rc_card(temp_path)
        
        # Clean up temp file
        try:
            os.remove(temp_path)
        except:
            pass
        
        if not result.get('success'):
            return jsonify({
                'success': False,
                'error': result.get('error', 'Failed to extract RC details')
            }), 400
        
        # Return extracted data
        extracted_data = result.get('data', {})
        
        return jsonify({
            'success': True,
            'data': {
                'rc': extracted_data.get('rc_number'),
                'owner': extracted_data.get('owner'),
                'bike': extracted_data.get('vehicle_model'),
                'chassis': extracted_data.get('chassis_number'),
                'engine': extracted_data.get('engine_number'),
                'fuel': extracted_data.get('fuel_type'),
                'class': extracted_data.get('class'),
                'received_on': extracted_data.get('registration_date')
            },
            'confidence': result.get('confidence', 0),
            'raw_text': result.get('raw_text', '')
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error processing RC card: {str(e)}'
        }), 500


# ================= DELETE VEHICLE API =================
@routes.route('/api/delete_vehicle/<int:id>', methods=['POST'])
@login_required
def delete_vehicle_api(id):
    """Delete a vehicle via API (returns JSON)"""
    if current_user.role != 'Admin':
        return jsonify({'success': False, 'error': 'Permission denied'}), 403
    
    try:
        v = Vehicle.query.get(id)
        if not v:
            return jsonify({'success': False, 'error': 'Vehicle not found'}), 404
        
        # Delete associated files
        if v.image and os.path.exists(os.path.join(UPLOAD_FOLDER, v.image)):
            os.remove(os.path.join(UPLOAD_FOLDER, v.image))
        
        if v.rc_card and os.path.exists(os.path.join(UPLOAD_FOLDER, v.rc_card)):
            os.remove(os.path.join(UPLOAD_FOLDER, v.rc_card))
        
        if v.insurance_file and os.path.exists(os.path.join(UPLOAD_FOLDER, v.insurance_file)):
            os.remove(os.path.join(UPLOAD_FOLDER, v.insurance_file))
        
        # Delete database records
        db.session.delete(v)
        db.session.commit()
        
        log_activity(0, f"Vehicle '{v.name}' deleted by {current_user.username}")
        
        return jsonify({
            'success': True,
            'message': f'Vehicle "{v.name}" deleted successfully'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Error deleting vehicle: {str(e)}'
        }), 500


# ================= DELETE USER API =================
@routes.route('/api/delete_user/<int:id>', methods=['POST'])
@login_required
def delete_user_api(id):
    """Delete a user via API (returns JSON)"""
    if current_user.role != 'Admin':
        return jsonify({'success': False, 'error': 'Permission denied'}), 403
    
    if current_user.id == id:
        return jsonify({'success': False, 'error': 'Cannot delete your own account'}), 400
    
    try:
        user = User.query.get(id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        db.session.delete(user)
        db.session.commit()
        
        log_activity(0, f"User '{user.username}' deleted by {current_user.username}")
        
        return jsonify({
            'success': True,
            'message': f'User "{user.username}" deleted successfully'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Error deleting user: {str(e)}'
        }), 500


# ================= DELETE VEHICLE IMAGE API =================
@routes.route('/api/delete_vehicle_image/<int:vehicle_id>', methods=['POST'])
@login_required
def delete_vehicle_image(vehicle_id):
    """Delete vehicle main image"""
    try:
        vehicle = Vehicle.query.get(vehicle_id)
        if not vehicle:
            return jsonify({'success': False, 'error': 'Vehicle not found'}), 404
        
        if current_user.role != 'Admin':
            return jsonify({'success': False, 'error': 'Permission denied'}), 403
        
        # Delete image file
        if vehicle.image and os.path.exists(os.path.join(UPLOAD_FOLDER, vehicle.image)):
            os.remove(os.path.join(UPLOAD_FOLDER, vehicle.image))
        
        # Update database
        vehicle.image = None
        db.session.commit()
        
        log_activity(vehicle_id, f"Vehicle image deleted by {current_user.username}")
        
        return jsonify({
            'success': True,
            'message': 'Image deleted successfully'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Error deleting image: {str(e)}'
        }), 500


# ================= DELETE RETURN IMAGE API =================
@routes.route('/api/delete_return_image/<int:vehicle_id>', methods=['POST'])
@login_required
def delete_return_image(vehicle_id):
    """Delete vehicle return image"""
    try:
        vehicle = Vehicle.query.get(vehicle_id)
        if not vehicle:
            return jsonify({'success': False, 'error': 'Vehicle not found'}), 404
        
        if current_user.role != 'Admin':
            return jsonify({'success': False, 'error': 'Permission denied'}), 403
        
        # Delete return image file
        if vehicle.return_image and os.path.exists(os.path.join(UPLOAD_FOLDER, vehicle.return_image)):
            os.remove(os.path.join(UPLOAD_FOLDER, vehicle.return_image))
        
        # Update database
        vehicle.return_image = None
        db.session.commit()
        
        log_activity(vehicle_id, f"Return image deleted by {current_user.username}")
        
        return jsonify({
            'success': True,
            'message': 'Return image deleted successfully'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Error deleting return image: {str(e)}'
        }), 500