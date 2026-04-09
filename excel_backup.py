"""
Excel Backup Utility - Automatically backup all vehicles, activities, and instrumentation
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
from models import Vehicle, Instrumentation, Activity

BACKUP_FOLDER = "backups"
BACKUP_FILE = os.path.join(BACKUP_FOLDER, "vehicles_backup.xlsx")

def ensure_backup_folder():
    """Create backup folder if it doesn't exist"""
    if not os.path.exists(BACKUP_FOLDER):
        os.makedirs(BACKUP_FOLDER)

def create_excel_header(ws):
    """Create header row with formatting"""
    headers = [
        'ID', 'Vehicle Name', 'RC Number', 'Owner', 'Bike Model', 'Chassis', 
        'Engine', 'Fuel Type', 'Status', 'Received On', 'Insurance', 
        'Number Plate', 'Registration Date', 'Manufacturing Date',
        'Inertia', 'CC', 'Gears', 'Road A', 'Road B', 'Road C',
        'Oil', 'Coolant', 'Start Date', 'End Date',
        'Keys', 'Scratches', 'Dents', 'Glass Damage', 'Tire Condition',
        'Battery Status', 'Lights Issue', 'Engine Status', 'Condition Notes',
        'Created At'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="B60000", end_color="B60000", fill_type="solid")  # Bosch Red
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def create_activities_header(ws):
    """Create header row for activities sheet"""
    headers = ['Activity ID', 'Vehicle ID', 'Vehicle Name', 'Description', 'Date', 'Performed By', 'Timestamp']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")  # Blue
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def create_instrumentation_header(ws):
    """Create header row for instrumentation sheet"""
    headers = ['Inst ID', 'Vehicle ID', 'Vehicle Name', 'Type', 'Activity', 'Request Date', 'Completion Date', 'Details']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")  # Green
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def backup_vehicles_to_excel(db):
    """
    Backup all vehicles from database to Excel file
    This function overwrites the entire Excel file with latest data
    """
    try:
        ensure_backup_folder()
        
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vehicles"
        
        # Create header
        create_excel_header(ws)
        
        # Fetch all vehicles
        vehicles = Vehicle.query.all()
        
        # Define column mapping
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add vehicle data
        for row, vehicle in enumerate(vehicles, 2):
            data = [
                vehicle.id,
                vehicle.name,
                vehicle.rc,
                vehicle.owner,
                vehicle.bike,
                vehicle.chassis,
                vehicle.engine,
                vehicle.fuel,
                vehicle.status,
                vehicle.received_on,
                vehicle.insurance,
                vehicle.number_plate,
                vehicle.reg_date,
                vehicle.mfg_date,
                vehicle.inertia,
                vehicle.cc,
                vehicle.gears,
                vehicle.road_a,
                vehicle.road_b,
                vehicle.road_c,
                vehicle.oil,
                vehicle.coolant,
                vehicle.start_date,
                vehicle.end_date,
                vehicle.number_of_keys,
                "Yes" if vehicle.scratches_present else "No",
                "Yes" if vehicle.dents_present else "No",
                "Yes" if vehicle.glass_damage_present else "No",
                vehicle.tire_condition,
                vehicle.battery_status,
                "Yes" if vehicle.lights_issue_present else "No",
                vehicle.engine_status,
                vehicle.condition_notes,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # Color alternate rows for readability
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Save workbook
        wb.save(BACKUP_FILE)
        print(f"✓ Vehicle backup saved: {BACKUP_FILE} ({len(vehicles)} vehicles)")
        return True
    
    except Exception as e:
        print(f"✗ Excel backup error: {str(e)}")
        return False

def add_vehicle_to_backup(vehicle, db):
    """
    Add a single vehicle to the backup Excel file
    This appends a new row if the file exists, or creates it if it doesn't
    """
    try:
        ensure_backup_folder()
        
        # If file exists, open it; otherwise create new workbook
        if os.path.exists(BACKUP_FILE):
            wb = openpyxl.load_workbook(BACKUP_FILE)
            ws = wb.active
            next_row = ws.max_row + 1
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vehicles"
            create_excel_header(ws)
            next_row = 2
        
        # Add vehicle data
        data = [
            vehicle.id,
            vehicle.name,
            vehicle.rc,
            vehicle.owner,
            vehicle.bike,
            vehicle.chassis,
            vehicle.engine,
            vehicle.fuel,
            vehicle.status,
            vehicle.received_on,
            vehicle.insurance,
            vehicle.number_plate,
            vehicle.reg_date,
            vehicle.mfg_date,
            vehicle.inertia,
            vehicle.cc,
            vehicle.gears,
            vehicle.road_a,
            vehicle.road_b,
            vehicle.road_c,
            vehicle.oil,
            vehicle.coolant,
            vehicle.start_date,
            vehicle.end_date,
            vehicle.number_of_keys,
            "Yes" if vehicle.scratches_present else "No",
            "Yes" if vehicle.dents_present else "No",
            "Yes" if vehicle.glass_damage_present else "No",
            vehicle.tire_condition,
            vehicle.battery_status,
            "Yes" if vehicle.lights_issue_present else "No",
            vehicle.engine_status,
            vehicle.condition_notes,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write data to row
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=next_row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Color alternate rows
            if next_row % 2 == 0:
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Save workbook
        wb.save(BACKUP_FILE)
        print(f"✓ Vehicle '{vehicle.name}' backed up to Excel")
        return True
    
    except Exception as e:
        print(f"✗ Excel backup error: {str(e)}")
        return False

def backup_all_data(db):
    """
    Complete backup of all data: Vehicles, Activities, and Instrumentation
    Creates/updates Excel file with 3 sheets
    """
    try:
        ensure_backup_folder()
        
        # Create new workbook with multiple sheets
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # ===== VEHICLES SHEET =====
        ws_vehicles = wb.create_sheet("Vehicles", 0)
        create_excel_header(ws_vehicles)
        
        vehicles = Vehicle.query.all()
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add vehicle data
        for row, vehicle in enumerate(vehicles, 2):
            data = [
                vehicle.id, vehicle.name, vehicle.rc, vehicle.owner, vehicle.bike,
                vehicle.chassis, vehicle.engine, vehicle.fuel, vehicle.status,
                vehicle.received_on, vehicle.insurance, vehicle.number_plate,
                vehicle.reg_date, vehicle.mfg_date, vehicle.inertia, vehicle.cc,
                vehicle.gears, vehicle.road_a, vehicle.road_b, vehicle.road_c,
                vehicle.oil, vehicle.coolant, vehicle.start_date, vehicle.end_date,
                vehicle.number_of_keys, "Yes" if vehicle.scratches_present else "No",
                "Yes" if vehicle.dents_present else "No",
                "Yes" if vehicle.glass_damage_present else "No", vehicle.tire_condition,
                vehicle.battery_status, "Yes" if vehicle.lights_issue_present else "No",
                vehicle.engine_status, vehicle.condition_notes,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_vehicles.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Auto-adjust column widths for vehicles
        for column in ws_vehicles.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_vehicles.column_dimensions[column_letter].width = min(max_length + 2, 50)
        ws_vehicles.row_dimensions[1].height = 30
        
        # ===== ACTIVITIES SHEET =====
        ws_activities = wb.create_sheet("Activities", 1)
        create_activities_header(ws_activities)
        
        activities = Activity.query.all()
        for row, activity in enumerate(activities, 2):
            # Get vehicle name
            vehicle = Vehicle.query.get(activity.vehicle_id) if activity.vehicle_id else None
            vehicle_name = vehicle.name if vehicle else "N/A"
            
            data = [
                activity.id, activity.vehicle_id, vehicle_name, activity.description,
                activity.date, activity.performed_by, activity.timestamp
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_activities.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        
        # Auto-adjust column widths for activities
        for column in ws_activities.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_activities.column_dimensions[column_letter].width = min(max_length + 2, 50)
        ws_activities.row_dimensions[1].height = 30
        
        # ===== INSTRUMENTATION SHEET =====
        ws_instrumentation = wb.create_sheet("Instrumentation", 2)
        create_instrumentation_header(ws_instrumentation)
        
        instrumentation = Instrumentation.query.all()
        for row, inst in enumerate(instrumentation, 2):
            # Get vehicle name
            vehicle = Vehicle.query.get(inst.vehicle_id) if inst.vehicle_id else None
            vehicle_name = vehicle.name if vehicle else "N/A"
            
            data = [
                inst.id, inst.vehicle_id, vehicle_name, inst.type, inst.activity,
                inst.request_date, inst.completion_date, inst.details
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_instrumentation.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        
        # Auto-adjust column widths for instrumentation
        for column in ws_instrumentation.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_instrumentation.column_dimensions[column_letter].width = min(max_length + 2, 50)
        ws_instrumentation.row_dimensions[1].height = 30
        
        # Save workbook
        wb.save(BACKUP_FILE)
        print(f"✓ Complete backup saved: {BACKUP_FILE} ({len(vehicles)} vehicles, {len(activities)} activities, {len(instrumentation)} instrumentation records)")
        return True
    
    except Exception as e:
        print(f"✗ Excel backup error: {str(e)}")
        return False

def auto_update_backup(db):
    """
    Auto-update the backup with the latest data
    Call this after every add/edit/delete operation
    """
    return backup_all_data(db)

def read_backup_data():
    """
    Read all data from the backup Excel file
    Returns list of dictionaries with vehicle data
    """
    try:
        if not os.path.exists(BACKUP_FILE):
            return []
        
        wb = openpyxl.load_workbook(BACKUP_FILE)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Read all vehicle data
        vehicles_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # Skip empty rows
                vehicle_dict = dict(zip(headers, row))
                vehicles_data.append(vehicle_dict)
        
        return vehicles_data
    
    except Exception as e:
        print(f"✗ Error reading backup: {str(e)}")
        return []

def restore_from_backup(db):
    """
    Restore all data (vehicles, activities, instrumentation) from backup Excel file
    Clears existing data and replaces with backup data
    """
    try:
        if not os.path.exists(BACKUP_FILE):
            return False, "Backup file not found"
        
        wb = openpyxl.load_workbook(BACKUP_FILE)
        
        # Clear existing data
        Vehicle.query.delete()
        Activity.query.delete()
        Instrumentation.query.delete()
        db.session.commit()
        
        restored_vehicles = 0
        restored_activities = 0
        restored_instrumentation = 0
        
        # ===== RESTORE VEHICLES =====
        if "Vehicles" in wb.sheetnames:
            ws_vehicles = wb["Vehicles"]
            for row in ws_vehicles.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:  # Skip empty rows
                    try:
                        vehicle = Vehicle(
                            name=row[1],  # Vehicle Name
                            rc=row[2],  # RC Number
                            owner=row[3],  # Owner
                            bike=row[4],  # Bike Model
                            chassis=row[5],  # Chassis
                            engine=row[6],  # Engine
                            fuel=row[7],  # Fuel Type
                            status=row[8] or 'In Testing',  # Status
                            received_on=row[9],  # Received On
                            insurance=row[10],  # Insurance
                            number_plate=row[11],  # Number Plate
                            reg_date=row[12],  # Registration Date
                            mfg_date=row[13],  # Manufacturing Date
                            inertia=row[14],  # Inertia
                            cc=row[15],  # CC
                            gears=row[16],  # Gears
                            road_a=row[17],  # Road A
                            road_b=row[18],  # Road B
                            road_c=row[19],  # Road C
                            oil=row[20],  # Oil
                            coolant=row[21],  # Coolant
                            start_date=row[22],  # Start Date
                            end_date=row[23],  # End Date
                            number_of_keys=row[24] or 1,  # Keys
                            scratches_present=row[25] == 'Yes',  # Scratches
                            dents_present=row[26] == 'Yes',  # Dents
                            glass_damage_present=row[27] == 'Yes',  # Glass Damage
                            tire_condition=row[28],  # Tire Condition
                            battery_status=row[29],  # Battery Status
                            lights_issue_present=row[30] == 'Yes',  # Lights Issue
                            engine_status=row[31],  # Engine Status
                            condition_notes=row[32]  # Condition Notes
                        )
                        db.session.add(vehicle)
                        restored_vehicles += 1
                    except Exception as e:
                        print(f"✗ Error restoring vehicle: {str(e)}")
                        continue
            db.session.commit()
        
        # ===== RESTORE ACTIVITIES =====
        if "Activities" in wb.sheetnames:
            ws_activities = wb["Activities"]
            for row in ws_activities.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:  # Skip empty rows
                    try:
                        activity = Activity(
                            id=row[0],  # Activity ID
                            vehicle_id=row[1],  # Vehicle ID
                            description=row[3],  # Description
                            date=row[4],  # Date
                            performed_by=row[5],  # Performed By
                            timestamp=row[6]  # Timestamp
                        )
                        db.session.add(activity)
                        restored_activities += 1
                    except Exception as e:
                        print(f"✗ Error restoring activity: {str(e)}")
                        continue
            db.session.commit()
        
        # ===== RESTORE INSTRUMENTATION =====
        if "Instrumentation" in wb.sheetnames:
            ws_instrumentation = wb["Instrumentation"]
            for row in ws_instrumentation.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:  # Skip empty rows
                    try:
                        inst = Instrumentation(
                            id=row[0],  # Inst ID
                            vehicle_id=row[1],  # Vehicle ID
                            type=row[3],  # Type
                            activity=row[4],  # Activity
                            request_date=row[5],  # Request Date
                            completion_date=row[6],  # Completion Date
                            details=row[7]  # Details
                        )
                        db.session.add(inst)
                        restored_instrumentation += 1
                    except Exception as e:
                        print(f"✗ Error restoring instrumentation: {str(e)}")
                        continue
            db.session.commit()
        
        message = f"✓ Restored {restored_vehicles} vehicles, {restored_activities} activities, {restored_instrumentation} instrumentation records"
        print(message)
        return True, message
    
    except Exception as e:
        print(f"✗ Error restoring backup: {str(e)}")
        return False, f"Error restoring backup: {str(e)}"
